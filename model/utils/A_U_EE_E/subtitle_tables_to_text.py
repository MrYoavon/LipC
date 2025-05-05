import os
import csv
from openpyxl import load_workbook


def time_to_milliseconds(hours, minutes, seconds, milliseconds):
    # Convert time to milliseconds
    return (int(hours) * 3600 + int(minutes) * 60 + int(seconds)) * 1000 + int(milliseconds)


def process_table(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, delimiter=',')

        # Assuming the first row is a header
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Assuming row format: [start_hms, start_ms, end_hms, end_ms, subtitle]
            # start time hours:minutes:seconds
            start_hms = str(row[0]).split(':')
            start_ms = str(row[1])  # start time milliseconds
            end_hms = str(row[2]).split(':')  # end time hours:minutes:seconds
            end_ms = str(row[3])  # end time milliseconds
            subtitle = row[4]  # subtitle text

            # Convert start and end times to milliseconds
            start_time = time_to_milliseconds(
                start_hms[0], start_hms[1], start_hms[2], start_ms)
            end_time = time_to_milliseconds(
                end_hms[0], end_hms[1], end_hms[2], end_ms)

            # Write the result to the CSV file
            writer.writerow([start_time, end_time, subtitle])


def process_all_tables(input_folder, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Loop over all files in the input folder
    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):  # Assuming the tables are in .xlsx format
            input_file = os.path.join(input_folder, filename)
            output_file = os.path.join(
                output_folder, filename.replace(".xlsx", ".csv"))
            process_table(input_file, output_file)
            print(f"Processed {filename}")


# Example usage:
input_folder = 'data/A_U_EE_E'
output_folder = 'data/A_U_EE_E/subtitles'

process_all_tables(input_folder, output_folder)
